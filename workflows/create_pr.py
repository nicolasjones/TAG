"""
Workflow: CREATE_PR
===================
Traduce el flujo Make.com "Calls from Bitrix24 - PR → Ruta CREATE_PR"
al framework Antigravity/TAG usando Microsoft 365 (SharePoint + openpyxl).

Flujo:
  1. Recibe payload de Bitrix24 con los datos del PR.
  2. Crea carpeta en SharePoint: '{SHAREPOINT_PR_ROOT}/{PR_Nombre}'.
  3. Descarga el template XLSX desde SharePoint.
  4. Llena la CARÁTULA con los datos del PR (placeholders {{...}}).
  5. Loop paginado: obtiene ítems del SPA "Apertura PR" en Bitrix24.
  6. Para cada ítem: escribe en la solapa "APERTURA DE PR" del Excel.
  7. Sube el Excel completo a la carpeta del PR en SharePoint.
  8. (Opcional) Crea tareas en Microsoft Planner si está configurado.

Mapping Make.com → Excel template (APERTURA DE PR):
  ┌─────────────┬────────┬────────────────────────────────────────────────────┐
  │ Columna     │ Make   │ Campo Bitrix / Valor                               │
  ├─────────────┼────────┼────────────────────────────────────────────────────┤
  │ A  (col  1) │ idx  0 │ Codif. concatenada = TIPO_INGENIERIA + " - " + E  │
  │ E  (col  5) │ idx  4 │ UF_CRM_7_1745325896 → CÓDIGO                      │
  │ G  (col  7) │ idx  6 │ UF_CRM_7_1747688168 → REVISIÓN DEL DOC.           │
  │ H  (col  8) │ idx  7 │ switch(UF_CRM_7_1747688217) → TIPO DE REVISION    │
  │             │        │   69→LETRA | 71→NÚMERO | 73→CO | 75→PID           │
  │ I  (col  9) │ idx  8 │ title → DESCRIPCIÓN DEL DOCUMENTO                 │
  │ J  (col 10) │ idx  9 │ UF_CRM_7_1747688282 → HORAS TOTALES               │
  │ K  (col 11) │ idx 10 │ UF_CRM_7_1747688577 → CANTIDAD DE HOJAS           │
  │ S  (col 19) │ idx 18 │ switch(UF_CRM_7_1747688317) → LINEA DEL CONTRATO  │
  │             │        │   77→ESTANDAR | 79→HORA ESPECIAL | 81→RASTREO M2  │
  │             │        │   83→DÍA DE RASTREO | 85→TRABAJO ESPECIAL         │
  │ V  (col 22) │ idx 21 │ UF_CRM_7_1747688406 → DÍAS PARA ENTREGA REV A    │
  └─────────────┴────────┴────────────────────────────────────────────────────┘

  Fila de headers: fila 4
  Primera fila de datos: fila 5
  Paginación: row = 4 + (page * 50) + item_index_1based

Placeholders del template (reemplazados con datos del payload):
  CARÁTULA:       E3={{NOMBRE_PR}}, H3={{REVISIÓN_PR}}, D4={{DENOMINACIÓN}},
                  E5={{NRO_DE_CAMBIO}}, H5={{LIDER_EMPRESA}}, E6={{COMPLEJO}},
                  H6={{LIDER_YPF}}, E7={{AREA}}, H7={{OT_PEP}}, E8={{UNIDAD}},
                  H8={{OTI_EMPRESA}}, E9={{FECHA_INI_PR}}, A12={{TIPO_INGENIERIA}},
                  A15={{DESCRIPCIÓN_TAREAS}}
  APERTURA DE PR: F1={{NOMBRE_PR}}, C3={{TIPO_INGENIERIA}}, W3={{PR_FECHA_APROB}}
                  A5-A1165={{TIPO_INGENIERIA}} - (prefijo de codificación)
"""

import io
import math
from datetime import datetime
from typing import Optional, List

from src.connectors.bitrix import BitrixConnector
from src.connectors.azure_custom import AzureCustomConnector
from src.core.config import settings
from src.core.logger import log_step, logger

from openpyxl import load_workbook
import plotly.express as px
import pandas as pd

# ─────────────────────────────────────────────────────────────────────────────
#  Constantes
# ─────────────────────────────────────────────────────────────────────────────

BITRIX_STAGE_APERTURA = "DT1040_13:UC_1QKUZC"
BITRIX_UF_PR_ID       = "UF_CRM_7_1744936769"
REPEATER_PAGES        = 10
ITEMS_PER_PAGE        = 50
APERTURA_HEADER_ROW   = 4    # Fila de títulos en APERTURA DE PR
APERTURA_DATA_START   = 5    # Primera fila de datos (= APERTURA_HEADER_ROW + 1)

# Switch TIPO DE REVISION (UF_CRM_7_1747688217) — columna H
TIPO_REVISION_MAP = {69: "LETRA", 71: "NÚMERO", 73: "CO", 75: "PID"}

# Switch LINEA DEL CONTRATO (UF_CRM_7_1747688317) — columna S (Debe coincidir EXACTO con CARÁTULA!I22:I26)
LINEA_CONTRATO_MAP = {
    77: "ESTANDAR",
    79: "HORA ESPECIAL",
    81: "RASTREO M2",
    83: "DIA DE RASTREO",
    85: "TRABAJO ESPECIAL (FACTURA)",
}

# Mapeo de Tipo de Ingeniería a los valores oficiales del template (columna A de lookup)
TIPO_INGENIERIA_EXCEL_MAP = {
    "BASICA":             "INGENIERÍA BÁSICA",
    "BÁSICA":             "INGENIERÍA BÁSICA",
    "DETALLE":            "INGENIERÍA DE DETALLE",
    "CONCEPTUAL":         "INGENIERÍA CONCEPTUAL",
    "BASICA EXTENDIDA":   "INGENIERÍA BÁSICA EXTENDIDA",
    "CONFORME A OBRA":    "INGENIERÍA CONFORME OBRA",
    "P&ID":               "P&ID",
}


# ─────────────────────────────────────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _parse_bitrix_date(date_str: str) -> Optional[datetime]:
    if not date_str: return None
    try:
        # Algunos campos de Bitrix vienen en formato ISO, otros solo fecha
        return datetime.fromisoformat(date_str.replace("Z", "+00:00"))
    except Exception:
        return None


def _switch(raw_value, mapping: dict, default: str = "") -> str:
    """Replica la función switch() de Make.com."""
    if raw_value is None:
        return default
    try:
        return mapping.get(int(raw_value), default)
    except (ValueError, TypeError):
        return default


def _safe_str(value) -> str:
    """Convierte a str manejando None."""
    return str(value) if value is not None else ""


# ─────────────────────────────────────────────────────────────────────────────
#  Llenado del template Excel
# ─────────────────────────────────────────────────────────────────────────────

def _fill_placeholders(wb, payload: dict) -> None:
    """
    Reemplaza todos los {{PLACEHOLDER}} del template en todas las hojas
    usando los datos del payload del PR.

    Placeholders soportados (extraídos del template NUEVO Template PR.xlsx):
      CARÁTULA:       {{NOMBRE_PR}}, {{REVISIÓN_PR}}, {{DENOMINACIÓN}},
                      {{NRO_DE_CAMBIO}}, {{LIDER_EMPRESA}}, {{COMPLEJO}},
                      {{LIDER_YPF}}, {{AREA}}, {{OT_PEP}}, {{UNIDAD}},
                      {{OTI_EMPRESA}}, {{FECHA_INI_PR}}, {{TIPO_INGENIERIA}},
                      {{DESCRIPCIÓN_TAREAS}}
      APERTURA DE PR: {{NOMBRE_PR}}, {{TIPO_INGENIERIA}}, {{PR_FECHA_APROB}}
                      (y el prefijo {{TIPO_INGENIERIA}} - en las filas de datos)
    """
    mapping = {
        "{{NOMBRE_PR}}":          payload.get("PR_Nombre", ""),
        "{{REVISIÓN_PR}}":        payload.get("PR_Revisión", ""),
        "{{DENOMINACIÓN}}":       payload.get("PR_Denominación", ""),
        "{{NRO_DE_CAMBIO}}":      payload.get("PR_NroDeCambio", ""),
        "{{LIDER_EMPRESA}}":      payload.get("PR_LiderEmpresa", ""),
        "{{COMPLEJO}}":           payload.get("PR_Complejo", ""),
        "{{LIDER_YPF}}":          payload.get("PR_LiderYPF", ""),
        "{{AREA}}":               payload.get("PR_Area", ""),
        "{{OT_PEP}}":             payload.get("PR_OT-PEP", ""),
        "{{UNIDAD}}":             payload.get("PR_Unidad", ""),
        "{{OTI_EMPRESA}}":        payload.get("PR_OTI-EMPRESA", ""),
        "{{FECHA_INI_PR}}":       payload.get("PR_FECHA-INI", ""),
        "{{TIPO_INGENIERIA}}":    payload.get("PR_TipoIngenieria", ""),
        "{{DESCRIPCIÓN_TAREAS}}": payload.get("PR_DescripciónTareas", ""),
        "{{PR_FECHA_APROB}}":     payload.get("PR_FECHA_APROB", ""),
    }

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "{{" in cell.value:
                    # Si el placeholder es lo único en la celda y el valor es un objeto (fecha), reemplazamos la celda entera
                    stripped_val = cell.value.strip()
                    if stripped_val in mapping and not isinstance(mapping[stripped_val], (str, int, float)):
                         cell.value = mapping[stripped_val]
                         continue
                    
                    # De lo contrario, reemplazo de string normal
                    new_val = cell.value
                    for placeholder, value in mapping.items():
                        new_val = new_val.replace(placeholder, _safe_str(value))
                    cell.value = new_val


def _fill_apertura_items(wb, items: List[dict], tipo_ingenieria: str) -> None:
    """
    Escribe los ítems de Bitrix en la solapa "APERTURA DE PR" del workbook.

    Estructura del template (fila 4 = headers, datos desde fila 5):
    ┌──────────────────────────────────────────────────────────────────┐
    │ Col A (1)  │ Codificación concatenada (TIPO - CÓDIGO)            │
    │ Col E (5)  │ CÓDIGO          ← UF_CRM_7_1745325896              │
    │ Col G (7)  │ REVISIÓN DOC.   ← UF_CRM_7_1747688168              │
    │ Col H (8)  │ TIPO DE REV.    ← switch(UF_CRM_7_1747688217)      │
    │ Col I (9)  │ DESCRIPCIÓN     ← title                            │
    │ Col J (10) │ HORAS TOTALES   ← UF_CRM_7_1747688282              │
    │ Col K (11) │ CANT. HOJAS     ← UF_CRM_7_1747688577              │
    │ Col S (19) │ LINEA CONTRATO  ← switch(UF_CRM_7_1747688317)      │
    │ Col V (22) │ DÍAS ENTREGA    ← UF_CRM_7_1747688406              │
    └──────────────────────────────────────────────────────────────────┘

    Paginación: row = 4 + (page * 50) + item_index_1based
    (replica el sum(4; 82.i * 50; 31.__IMTINDEX__) de Make.com)
    """
    ws = wb["APERTURA DE PR"]

    for global_idx, item in enumerate(items, start=1):
        # Calcular fila exacta (mismo cálculo que Make.com)
        page  = (global_idx - 1) // ITEMS_PER_PAGE
        idx_in_page = ((global_idx - 1) % ITEMS_PER_PAGE) + 1
        row = APERTURA_HEADER_ROW + (page * ITEMS_PER_PAGE) + idx_in_page

        # Extraer campos del ítem Bitrix
        codigo      = _safe_str(item.get("UF_CRM_7_1745325896", ""))
        revision    = _safe_str(item.get("UF_CRM_7_1747688168", ""))
        tipo_rev    = _switch(item.get("UF_CRM_7_1747688217"), TIPO_REVISION_MAP)
        descripcion = _safe_str(item.get("title", ""))
        horas       = item.get("UF_CRM_7_1747688282", 0) or 0
        hojas       = item.get("UF_CRM_7_1747688577", 0) or 0
        linea       = _switch(item.get("UF_CRM_7_1747688317"), LINEA_CONTRATO_MAP)
        dias_entrega = item.get("UF_CRM_7_1747688406", "")

        # Col A: Codificación concatenada (TIPO_INGENIERIA - CÓDIGO)
        ws.cell(row=row, column=1).value  = f"{tipo_ingenieria} - {codigo}" if codigo else tipo_ingenieria

        # Col E: Código del documento
        ws.cell(row=row, column=5).value  = codigo

        # Col G: Revisión del documento (ej. "A", "1", "B")
        ws.cell(row=row, column=7).value  = revision

        # Col H: Tipo de revisión (LETRA / NÚMERO / CO / PID)
        ws.cell(row=row, column=8).value  = tipo_rev

        # Col I: Descripción del documento
        ws.cell(row=row, column=9).value  = descripcion

        # Col J: Horas totales presupuestadas
        ws.cell(row=row, column=10).value = horas

        # Col K: Cantidad de hojas del documento
        ws.cell(row=row, column=11).value = hojas

        # Col S: Línea del contrato (ESTANDAR / HORA ESPECIAL / etc.)
        ws.cell(row=row, column=19).value = linea

        # Col V: Días para entrega de Rev A una vez aprobado el PR
        ws.cell(row=row, column=22).value = _safe_str(dias_entrega)


def _generate_gantt_pdf(items: List[dict], pr_nombre: str) -> bytes:
    """Genera un gráfico de Gantt de alta calidad usando Plotly."""
    if not items:
        return None
        
    data = []
    for item in items:
        raw_start = item.get("begindate")
        raw_finish = item.get("closedate")
        start = _parse_bitrix_date(raw_start)
        finish = _parse_bitrix_date(raw_finish)
        
        if not start or not finish:
            continue
            
        # Si la fecha de fin es igual a la de inicio, sumamos 1 día para que sea visible
        if start == finish:
            from datetime import timedelta
            finish = finish + timedelta(days=1)
            
        data.append({
            "Tarea": f"{item.get('UF_CRM_7_1745325896', '')} - {item.get('title')}",
            "Inicio": start,
            "Fin": finish,
            "Tipo": TIPO_REVISION_MAP.get(int(item.get("UF_CRM_7_1747688217") or 0), "Tarea")
        })

    if not data:
        return None

    df = pd.DataFrame(data)
    
    fig = px.timeline(
        df, 
        x_start="Inicio", 
        x_end="Fin", 
        y="Tarea", 
        color="Tipo",
        title=f"Cronograma - {pr_nombre}",
        template="plotly_white",
        color_discrete_sequence=px.colors.qualitative.Prism
    )
    
    fig.update_yaxes(autorange="reversed")
    fig.update_layout(
        font=dict(family="Arial", size=10),
        title_font=dict(size=16),
        margin=dict(l=50, r=50, t=80, b=50)
    )

    try:
        return fig.to_image(format="pdf", engine="kaleido", width=1200, height=800)
    except Exception as e:
        logger.error(f"Error generando Gantt PDF: {str(e)}")
        return None


# ─────────────────────────────────────────────────────────────────────────────
#  Flujo principal
# ─────────────────────────────────────────────────────────────────────────────

async def main(payload: dict = None):
    """
    Punto de entrada del workflow CREATE_PR.

    Payload esperado (enviado por Bitrix24 vía webhook):
    {
        "Function":              "CREATE_PR",
        "PR_Nombre":             "PR2025-0001",
        "PR_ID":                 "123",
        "PR_Area":               "CIEP - ALQUILACION",
        "PR_OT-PEP":             "OT-0001",
        "PR_Unidad":             "CIEP",
        "PR_Complejo":           "CILP",
        "PR_LiderYPF":           "Nombre Apellido",
        "PR_OTI-EMPRESA":        "OTI-2025-001",
        "PR_FECHA-INI":          "01/04/2025",
        "PR_Revisión":           "0",
        "PR_Denominación":       "Ampliación Destilación",
        "PR_LiderEmpresa":       "Nombre Apellido",
        "PR_NroDeCambio":        "NCT-001",
        "PR_TipoIngenieria":     "NUEVA INSTALACION",
        "PR_DescripciónTareas":  "Descripción...",
        "PR_FECHA_APROB":        "15/04/2025"
    }
    """
    if not payload:
        raise ValueError("Payload vacío — se requieren los datos del PR.")

    pr_nombre       = _safe_str(payload.get("PR_Nombre", "")).strip()
    
    if not pr_nombre:
        raise ValueError("Falta el campo 'PR_Nombre' (ej: PR2025-0069) en el payload.")

    log_step("CREATE_PR ── Iniciando búsqueda por nombre", {"PR_Nombre": pr_nombre})

    bitrix = BitrixConnector()

    # ── PASO 0: Identificar ID por Nombre y Deep Fetch ───────────────────
    log_step(f"PASO 0: Buscando PR '{pr_nombre}' en Bitrix (SPA 1096)")
    find_resp = await bitrix.execute("crm.item.list", {
        "entityTypeId": 1096,
        "filter": {"title": pr_nombre},
        "select": ["id", "title"]
    })
    
    items = find_resp.get("result", {}).get("items", [])
    if not items:
        raise ValueError(f"No se encontró ningún PR con el título '{pr_nombre}' en Bitrix.")
    
    pr_id = items[0]["id"]
    log_step(f"PR identificado con ID: {pr_id}")

    pr_resp = await bitrix.execute("crm.item.get", {"entityTypeId": 1096, "id": pr_id})
    pr_data = pr_resp.get("result", {}).get("item", {})
    
    if not pr_data:
        raise ValueError(f"Error al recuperar datos para el PR ID {pr_id}.")

    # Mapeo de campos de Bitrix a nuestro payload interno
    # (Usamos los campos reales identificados en la instancia del cliente)

    # 1. Numero de Cambio -> Numero de MOC (ufCrm33_1747751597)
    nro_cambio = _safe_str(pr_data.get("ufCrm33_1747751597", "")).strip()

    # 2. Complejo (ufCrm33_1747675939) -> Mapeo Enum
    complejo_id = pr_data.get("ufCrm33_1747675939")
    COMPLEJO_MAP = {"277": "RLP", "279": "QLP", "45": "CIEP", "47": "CILE", "49": "CILP", "53": "CIPH", "55": "YTEC"}
    complejo_val = COMPLEJO_MAP.get(str(complejo_id), "QLP") # Fallback to QLP as requested for this PR

    # 3. Preparar fechas (Usamos solo fecha sin hora para VLOOKUPs en Excel)
    def _to_date_obj(d_str):
        if not d_str: return None
        dt = _parse_bitrix_date(d_str)
        # Truncar a medianoche (date) para matching exacto en Excel
        return dt.date() if dt else None

    fetched_payload = {
        "PR_Nombre":             pr_data.get("title", pr_nombre),
        "PR_Denominación":       _safe_str(pr_data.get("ufCrm33_1747675458", "")).strip(),
        "PR_OT-PEP":             _safe_str(pr_data.get("ufCrm33_1747676129", "")).strip(),
        "PR_Revisión":           _safe_str(pr_data.get("ufCrm33_1747675500", "0")),
        "PR_NroDeCambio":        nro_cambio,
        "PR_Complejo":           complejo_val,
        "PR_OTI-EMPRESA":        _safe_str(pr_data.get("ufCrm33_1747676167", "")).strip(),
        "PR_DescripciónTareas":  _safe_str(pr_data.get("ufCrm33_1747676308", "")).strip(),
        "PR_FECHA-INI":          _to_date_obj(pr_data.get("begindate")),
        "PR_FECHA_APROB":        _to_date_obj(pr_data.get("ufCrm33_1764852413")),
    }

    # Lider YPF (Contacto vinculado)
    lider_ypf_raw = pr_data.get("ufCrm33_1747676478")
    if lider_ypf_raw:
        contact_id = str(lider_ypf_raw).replace("C_", "")
        contact_resp = await bitrix.execute("crm.contact.get", {"id": contact_id})
        c = contact_resp.get("result", {})
        fetched_payload["PR_LiderYPF"] = f"{c.get('NAME', '')} {c.get('LAST_NAME', '')}".strip()
    
    # Lider Empresa (Asignado)
    # Por defecto 'Nestor Cruz' si el ID es 19, o el nombre del asignado si permitiera user.get
    if pr_data.get("assignedById") == 19:
        fetched_payload["PR_LiderEmpresa"] = "Nestor Cruz"
    else:
        fetched_payload["PR_LiderEmpresa"] = "Nestor Cruz" # Fallback solicitado

    # Area (SPA 1110)
    area_id = pr_data.get("ufCrm33_1764164000")
    if area_id:
        area_resp = await bitrix.execute("crm.item.get", {"entityTypeId": 1110, "id": area_id})
        fetched_payload["PR_Area"] = area_resp.get("result", {}).get("item", {}).get("title", "AROMATICOS")
    
    # Unidad (SPA 1114)
    unidad_id = pr_data.get("ufCrm33_1764164126")
    if unidad_id:
        unidad_resp = await bitrix.execute("crm.item.get", {"entityTypeId": 1114, "id": unidad_id})
        fetched_payload["PR_Unidad"] = unidad_resp.get("result", {}).get("item", {}).get("title", "")

    # Tipo de Ingeniería (Mapeado)
    tipo_id = pr_data.get("ufCrm33_1747676203")
    # Mapping Bitrix Enum IDs -> Texto legible
    TIPO_ID_MAP = {"57": "BASICA", "61": "CONCEPTUAL", "63": "DETALLE", "59": "BASICA EXTENDIDA", "65": "CONFORME A OBRA"}
    raw_tipo = TIPO_ID_MAP.get(str(tipo_id), "BASICA")
    tipo_ingenieria = TIPO_INGENIERIA_EXCEL_MAP.get(raw_tipo, "INGENIERÍA BÁSICA")
    fetched_payload["PR_TipoIngenieria"] = tipo_ingenieria

    # Mezclamos con el payload original por si vienen datos extra
    log_step("Datos recuperados de Bitrix", fetched_payload)
    payload.update(fetched_payload)
    
    # ── PASO 1: Preparar SharePoint ───────────────────────────────────────
    azure  = AzureCustomConnector()
    drive_id      = settings.SHAREPOINT_PR_DRIVE_ID
    pr_root       = settings.SHAREPOINT_PR_ROOT
    template_path = settings.SHAREPOINT_TEMPLATE_PR_PATH

    # Sanitizar nombre para rutas de SharePoint
    pr_nombre_safe = pr_nombre.replace(":", "-").replace("/", "-").replace("\\", "-")

    log_step(f"PASO 1: Preparando carpeta '{pr_nombre_safe}' en SharePoint ({pr_root})")
    await azure.create_folder(drive_id, pr_root, pr_nombre_safe)
    folder_path = f"{pr_root}/{pr_nombre_safe}"
    log_step(f"Carpeta lista: {folder_path}")

    # ── PASO 2: Descargar template XLSX desde SharePoint ──────────────────
    log_step(f"PASO 2: Descargando template desde {template_path}")
    template_bytes = await azure.download_file(drive_id, template_path)
    log_step(f"Template descargado ({len(template_bytes):,} bytes)")

    # ── PASO 3: Abrir template y llenar placeholders de la CARÁTULA ───────
    log_step("PASO 3: Llenando placeholders del template (CARÁTULA + APERTURA DE PR)")
    wb = load_workbook(io.BytesIO(template_bytes))
    _fill_placeholders(wb, payload)

    # ── PASO 4: Loop paginado — obtener ítems de Bitrix ───────────────────
    # Replica: BasicRepeater (10 páginas) → crm.item.list → BasicFeeder
    log_step("PASO 4: Iniciando loop paginado de ítems Bitrix")
    all_items = []

    for page in range(REPEATER_PAGES):
        start = page * ITEMS_PER_PAGE
        log_step(f"  Página {page} (start={start})")

        resp = await bitrix.execute("crm.item.list", {
            "entityTypeId":     settings.BITRIX_ENTITY_APERTURA_PR,
            "useOriginalUfNames": "Y",
            "filter": {
                BITRIX_UF_PR_ID:  pr_id,
                "stageId":        BITRIX_STAGE_APERTURA,
            },
            "select": ["*"],
            "start": start,
        })

        items = resp.get("result", {}).get("items", [])
        log_step(f"  Ítems recibidos: {len(items)}")

        if not items:
            log_step(f"  Página {page} vacía — deteniendo loop")
            break

        all_items.extend(items)

    log_step(f"Total ítems de Bitrix: {len(all_items)}")

    # ── PASO 5: Llenar la solapa APERTURA DE PR con los ítems ─────────────
    log_step("PASO 5: Escribiendo ítems en la solapa 'APERTURA DE PR'")
    _fill_apertura_items(wb, all_items, tipo_ingenieria)

    # ── PASO 6: Serializar workbook y subir a SharePoint ──────────────────
    log_step("PASO 6: Guardando Excel y subiendo a SharePoint")
    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    excel_bytes = output_buffer.getvalue()

    filename = f"{pr_nombre_safe}.xlsx"
    await azure.upload_file(drive_id, folder_path, filename, excel_bytes)
    log_step(f"Excel subido: {folder_path}/{filename}")

    # ── PASO 7: Generar Gantt PDF y subir ─────────────────────────────────
    log_step("PASO 7: Generando Gantt PDF")
    gantt_bytes = _generate_gantt_pdf(all_items, pr_nombre)
    if gantt_bytes:
        gantt_filename = f"Gantt_{pr_nombre_safe}.pdf"
        await azure.upload_file(drive_id, folder_path, gantt_filename, gantt_bytes)
        log_step(f"Gantt PDF subido: {folder_path}/{gantt_filename}")
    else:
        log_step("Gantt omitido (sin fechas o sin ítems)")

    # ── PASO 8 (Opcional): Crear tareas en Microsoft Planner ──────────────
    plan_id   = getattr(settings, "PLANNER_PLAN_ID", None)
    bucket_id = getattr(settings, "PLANNER_BUCKET_ID", None)

    if plan_id and bucket_id and all_items:
        log_step(f"PASO 8: Creando tareas en Planner")
        for item in all_items:
            codigo = _safe_str(item.get("UF_CRM_7_1745325896", ""))
            title  = _safe_str(item.get("title", ""))
            await azure.create_planner_task(
                plan_id=plan_id,
                bucket_id=bucket_id,
                title=f"{codigo} - {title}" if codigo else title,
            )
        log_step("Tareas de Planner creadas")
    else:
        log_step("PASO 7: Planner no configurado — omitido")

    # ─────────────────────────────────────────────────────────────────────
    log_step("CREATE_PR ── Completado exitosamente", {
        "pr_nombre":    pr_nombre,
        "items":        len(all_items),
        "sharepoint":   f"{folder_path}/{filename}",
    })

    return {
        "status":             "completed",
        "pr_nombre":          pr_nombre,
        "sharepoint_folder":  folder_path,
        "archivo_excel":      filename,
        "total_items":        len(all_items),
    }
